import os
import openpyxl

# Tek seferlik ve excel to shared klasörlerinin yolları
tek_seferlik_arac_klasor_yolu = "/Users/seldacetin/Desktop/tekseferlik"
excel_to_shared_klasor_yolu= "/Users/seldacetin/Desktop/exceltoshared"

# AcilTirArac excel dosyasının yolu
acil_tir_arac_yolu = os.path.join(tek_seferlik_arac_klasor_yolu, 'AcilTirArac.xlsx')

# TekSeferlikArac2020 excel dosyasının yolu
tek_seferlik_arac_2020_yolu = os.path.join(excel_to_shared_klasor_yolu, 'TEK SEFERLİK ARAÇ 2020.xlsx')

# AcilTirArac excel dosyasını aç
acil_tir_arac = openpyxl.load_workbook(acil_tir_arac_yolu)
acil_tir_arac_sayfa = acil_tir_arac.active

# Verileri kopyala
veriler = []
for satir in acil_tir_arac_sayfa.iter_rows(min_row=2, max_row=acil_tir_arac_sayfa.max_row, min_col=2, max_col=12, values_only=True):
    veriler.extend(satir)

# TekSeferlikArac2020 excel dosyasını aç
tek_seferlik_arac_2020 = openpyxl.load_workbook(tek_seferlik_arac_2020_yolu)
tek_seferlik_arac_2020_sayfa = tek_seferlik_arac_2020['2020']

# B sütunundaki son verinin hücre adresini bul
b_sutun_son_veri_hucre_adresi = None
for satir in tek_seferlik_arac_2020_sayfa.iter_rows(min_row=1, max_row=tek_seferlik_arac_2020_sayfa.max_row, min_col=2, max_col=2):
    for hucre in satir:
        if hucre.value is None:
            b_sutun_son_veri_hucre_adresi = hucre.coordinate
            break
    if b_sutun_son_veri_hucre_adresi:
        break

# Verileri B sütunundaki son verinin hemen sonrasına yapıştır (B ve L sütunları arasına)
for i, deger in enumerate(veriler, start=2):
    tek_seferlik_arac_2020_sayfa.cell(row=i, column=2).value = deger  # B sütunu
    tek_seferlik_arac_2020_sayfa.cell(row=i, column=12).value = deger  # L sütunu

# Değişiklikleri kaydet
tek_seferlik_arac_2020.save(tek_seferlik_arac_2020_yolu)
